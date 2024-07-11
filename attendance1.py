import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Load the Excel workbook and worksheet
wb = openpyxl.load_workbook('attendance.xlsx')
ws = wb.active

# Get the next available column number
next_col = ws.max_column + 1

# Take attendance for each student
date = datetime.date.today()
print(f"Attendance for {date}:")
ws.cell(row=1, column=next_col, value=date)

def take_attendance(row_idx, row):
    reg_num, name = row[:2]
    while True:
        attendance = input(f"Enter attendance for {name} ({reg_num}): (p)resent or (a)bsent: ")
        if attendance.lower() == "p":
            ws.cell(row=row_idx + 2, column=next_col, value=1)
            if ws.cell(row=row_idx + 2, column=3).value is None:
                ws.cell(row=row_idx + 2, column=3, value=1)
            else:
                ws.cell(row=row_idx + 2, column=3, value=ws.cell(row=row_idx + 2, column=3).value + 1)
            if ws.cell(row=row_idx + 2, column=4).value is None:
                ws.cell(row=row_idx + 2, column=4, value=1)
            else:
                ws.cell(row=row_idx + 2, column=4, value=ws.cell(row=row_idx + 2, column=4).value + 1)
            break
        elif attendance.lower() == "a":
            ws.cell(row=row_idx + 2, column=next_col, value=0)
            if ws.cell(row=row_idx + 2, column=3).value is None:
                ws.cell(row=row_idx + 2, column=3, value=1)
            else:
                ws.cell(row=row_idx + 2, column=3, value=ws.cell(row=row_idx + 2, column=3).value + 1)
            break
        else:
            print("Invalid input. Please enter 'p' or 'a'.")
            take_attendance(row_idx, row)

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    take_attendance(row_idx, row)

# Save the Excel workbook
wb.save('attendance.xlsx')

print("\nAttendance updated successfully!")